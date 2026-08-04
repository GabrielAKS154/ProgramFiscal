import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import xml.etree.ElementTree as ET
import openpyxl
import os
import re
import unicodedata

BG      = "#1e1e2e"
SURFACE = "#2a2a3d"
ACCENT  = "#7c6af7"
ACCENT2 = "#5a4fcf"
TEXT    = "#e0e0f0"
MUTED   = "#8888aa"
SUCCESS = "#4caf82"   # verde
ERROR   = "#f06060"   # vermelho
WARNING = "#f0b060"   # laranja
INFO    = "#6ab0f5"   # azul claro
GRAY    = "#aaaaaa"   # cinza
BORDER  = "#3a3a55"

# Colunas fixas de resultado (ordem exata)
COLUNAS = [
    "CNPJ",
    "TIPO",
    "Art. LC 214/2025",
    "NCM/NBS",
    "DESCRICAO",
    "CST",
    "CCLASSTRIB ou CCREDPRES",
]

# Mapeamento coluna -> tags XML
MAPA_ITENS = {
    "CNPJ":                    ["emit/CNPJ", "CNPJ"],
    "TIPO":                    ["det/prod/xProd", "det/serv/xDescServ", "xProd", "xDescServ"],
    "ART. LC 214/2025":        [],
    "NCM/NBS":                 ["NCM", "NBS", "det/prod/NCM", "det/serv/NBS"],
    "DESCRICAO":               ["xProd", "xDescServ", "det/prod/xProd", "det/serv/xDescServ"],
    "CST":                     ["CST", "CSOSN",
                                "det/imposto/ICMS/ICMS00/CST",
                                "det/imposto/ICMS/ICMSSN102/CSOSN",
                                "det/imposto/PIS/PISAliq/CST",
                                "det/imposto/COFINS/COFINSAliq/CST"],
    "CCLASSTRIB OU CCREDPRES": ["cClassTrib", "cCredPres",
                                "det/imposto/cClassTrib", "det/imposto/cCredPres"],
}

# status -> tag de cor
STATUS_TAG = {
    "OK":                   "ok",
    "OK (raiz)":            "ok",
    "Divergente":           "erro",
    "Campo vazio no XML":   "vazio_xml",
    "Campo vazio no Excel": "vazio_excel",
    "Campo vazio":          "vazio_ambos",
    "Referencia legal":     "info",
}

def normaliza_cnpj(v):
    return re.sub(r"\D", "", str(v or ""))

def cnpj_raiz(v):
    return normaliza_cnpj(v)[:8]

def normaliza(v):
    s = str(v or "").strip().upper()
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")

def vazio(v):
    return v is None or str(v).strip() == ""


class XMLValidatorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("XML Validator  v2.6")
        self.geometry("1400x720")
        self.minsize(1000, 580)
        self.configure(bg=BG)
        self.excel_path       = tk.StringVar(value="Nenhum arquivo selecionado")
        self._excel_full_path = None
        self.xml_paths        = []
        self.results          = []
        self._build_ui()

    # ── UI ────────────────────────────────────────────────────────────────────
    def _build_ui(self):
        hdr = tk.Frame(self, bg=ACCENT, height=56)
        hdr.pack(fill="x")
        tk.Label(hdr, text="XML Validator", font=("Segoe UI", 16, "bold"),
                 bg=ACCENT, fg="white").pack(side="left", padx=20, pady=12)
        tk.Label(hdr, text="Validacao Fiscal NF-e  |  Reforma Tributaria  |  100% offline",
                 font=("Segoe UI", 9), bg=ACCENT, fg="#d0ccff").pack(side="left")

        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=20, pady=16)

        left = tk.Frame(body, bg=BG, width=280)
        left.pack(side="left", fill="y", padx=(0, 16))
        left.pack_propagate(False)

        self._section(left, "1. Formulario Excel (gabarito)")
        self._btn(left, "Selecionar Excel", self._load_excel).pack(fill="x", pady=(0, 4))
        tk.Label(left, textvariable=self.excel_path, bg=BG, fg=MUTED,
                 font=("Segoe UI", 8), wraplength=260, justify="left").pack(anchor="w")

        self._sep(left)
        self._section(left, "2. Arquivos XML (notas fiscais)")
        self._btn(left, "Adicionar XMLs", self._load_xmls).pack(fill="x", pady=(0, 4))
        self.xml_listbox = tk.Listbox(left, bg=SURFACE, fg=TEXT, selectbackground=ACCENT,
                                      relief="flat", bd=0, font=("Segoe UI", 8),
                                      height=7, highlightthickness=1,
                                      highlightbackground=BORDER)
        self.xml_listbox.pack(fill="x", pady=(0, 4))
        self._btn(left, "Remover selecionado", self._remove_xml, color=SURFACE).pack(fill="x")

        self._sep(left)
        self._section(left, "3. Executar")
        self._btn(left, "Validar XMLs", self._run, color=ACCENT).pack(fill="x", pady=(0, 4))
        self._btn(left, "Exportar relatorio", self._export, color=ACCENT2).pack(fill="x", pady=(0, 4))
        self._btn(left, "Limpar tudo", self._clear, color=SURFACE).pack(fill="x")

        self._sep(left)
        self._section(left, "Legenda (por linha)")
         for cor, txt in [
            (SUCCESS, "OK"),
            (ERROR,   "Divergente"),
            (WARNING, "Vazio no XML"),
            (INFO,    "Vazio no Excel"),
            (GRAY,    "Vazio (ambos)"),
        ]:
            row = tk.Frame(left, bg=BG)
            row.pack(anchor="w", pady=2)
            tk.Frame(row, bg=cor, width=14, height=14).pack(side="left", padx=(0, 8))
            tk.Label(row, text=txt, bg=BG, fg=cor,
                     font=("Segoe UI", 9, "bold")).pack(side="left")

        right = tk.Frame(body, bg=BG)
        right.pack(side="left", fill="both", expand=True)

        # Treeview colunas fixas
        col_ids = ["arquivo"] + [f"c{i}" for i in range(len(COLUNAS))]
        hdrs    = ["Arquivo XML"] + COLUNAS
        wids    = [200] + [150] * len(COLUNAS)

        self.tree = ttk.Treeview(right, columns=col_ids, show="headings", height=25)
        for cid, h, w in zip(col_ids, hdrs, wids):
            self.tree.heading(cid, text=h)
            self.tree.column(cid, width=w, minwidth=60, anchor="center")

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                        background=SURFACE, foreground=TEXT,
                        fieldbackground=SURFACE, rowheight=30,
                        font=("Segoe UI", 9))
        style.configure("Treeview.Heading",
                        background=BORDER, foreground=TEXT,
                        font=("Segoe UI", 9, "bold"))
                # --- Correção do bug que faz o Treeview ignorar as cores das tags ---
        def fixed_map(option):
            return [e for e in style.map("Treeview", query_opt=option)
                    if e[:2] != ("!disabled", "!selected")]

        style.map("Treeview",
                  foreground=fixed_map("foreground"),
                  background=fixed_map("background"),
                  fieldbackground=[("selected", ACCENT)])

        # tags — cada linha recebe a cor do pior status
        self.tree.tag_configure("ok",          foreground=SUCCESS, background="#1e2e24")
        self.tree.tag_configure("erro",        foreground=ERROR,   background="#2e1e1e")
        self.tree.tag_configure("vazio_xml",   foreground=WARNING, background="#2e2718")
        self.tree.tag_configure("vazio_excel", foreground=INFO,    background="#1a2430")
        self.tree.tag_configure("vazio_ambos", foreground=GRAY,    background="#232323")
        self.tree.tag_configure("info",        foreground=MUTED,   background=SURFACE)
        self.tree.tag_configure("misto",       foreground=TEXT,    background=SURFACE)

        sb_x = ttk.Scrollbar(right, orient="horizontal", command=self.tree.xview)
        sb_y = ttk.Scrollbar(right, orient="vertical",   command=self.tree.yview)
        self.tree.configure(xscrollcommand=sb_x.set, yscrollcommand=sb_y.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb_y.pack(side="left",  fill="y")
        sb_x.pack(side="bottom", fill="x")

        self.status_var = tk.StringVar(value="Pronto. Selecione o Excel e os XMLs para comecar.")
        tk.Label(self, textvariable=self.status_var, bg=SURFACE, fg=MUTED,
                 font=("Segoe UI", 8), anchor="w", padx=12, pady=6).pack(fill="x", side="bottom")

    def _section(self, parent, text):
        tk.Label(parent, text=text, bg=BG, fg=ACCENT,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(10, 2))

    def _sep(self, parent):
        tk.Frame(parent, bg=BORDER, height=1).pack(fill="x", pady=8)

    def _btn(self, parent, text, cmd, color=ACCENT):
        return tk.Button(parent, text=text, command=cmd, bg=color, fg="white",
                         activebackground=ACCENT2, activeforeground="white",
                         relief="flat", bd=0, font=("Segoe UI", 9),
                         cursor="hand2", pady=8)

    # ── Arquivos ──────────────────────────────────────────────────────────────
    def _load_excel(self):
        path = filedialog.askopenfilename(
            title="Selecionar formulario Excel",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")])
        if path:
            self.excel_path.set(os.path.basename(path))
            self._excel_full_path = path
            self.status_var.set(f"Excel carregado: {os.path.basename(path)}")

    def _load_xmls(self):
        paths = filedialog.askopenfilenames(
            title="Selecionar arquivos XML",
            filetypes=[("XML", "*.xml"), ("Todos", "*.*")])
        for p in paths:
            if p not in self.xml_paths:
                self.xml_paths.append(p)
                self.xml_listbox.insert("end", os.path.basename(p))
        self.status_var.set(f"{len(self.xml_paths)} arquivo(s) XML carregado(s).")

    def _remove_xml(self):
        sel = self.xml_listbox.curselection()
        if sel:
            idx = sel[0]
            self.xml_listbox.delete(idx)
            self.xml_paths.pop(idx)

    def _clear(self):
        self.excel_path.set("Nenhum arquivo selecionado")
        self._excel_full_path = None
        self.xml_paths.clear()
        self.xml_listbox.delete(0, "end")
        self.tree.delete(*self.tree.get_children())
        self.results.clear()
        self.status_var.set("Limpo. Pronto para nova validacao.")

    # ── Validacao ─────────────────────────────────────────────────────────────
    def _run(self):
        try:
            self._run_inner()
        except Exception as e:
            import traceback
            messagebox.showerror("Erro inesperado", traceback.format_exc())

    def _run_inner(self):
        if not self._excel_full_path:
            messagebox.showwarning("Atencao", "Selecione o formulario Excel primeiro.")
            return
        if not self.xml_paths:
            messagebox.showwarning("Atencao", "Adicione pelo menos um arquivo XML.")
            return

        self.tree.delete(*self.tree.get_children())
        self.results.clear()
        self.status_var.set("Processando...")
        self.update_idletasks()

        try:
            cnpj_excel, dados_tabela2 = self._read_excel(self._excel_full_path)
        except Exception as e:
            messagebox.showerror("Erro ao ler Excel", str(e))
            return

        ok = err = vz_xml = vz_excel = vz_ambos = info_cnt = 0

        # prioridade para determinar a cor da linha inteira
        PRIORIDADE = ["erro", "vazio_xml", "vazio_excel", "vazio_ambos", "info", "ok", "misto"]

        for xml_path in self.xml_paths:
            fname = os.path.basename(xml_path)
            try:
                xml_flat   = self._read_xml_flat(xml_path)
                excel_row  = dados_tabela2[0] if dados_tabela2 else {}

                row_valores  = [fname]
                row_tags     = []   # tag de cor de cada celula

                for col in COLUNAS:
                    col_norm = normaliza(col)

                    # busca tags XML para a coluna
                    tags_xml = None
                    for k in MAPA_ITENS:
                        if normaliza(k) == col_norm:
                            tags_xml = MAPA_ITENS[k]
                            break
                    if tags_xml is None:
                        tags_xml = [col]

                    # valor do Excel
                    if col_norm == "CNPJ":
                        val_excel = cnpj_excel
                    elif "ART" in col_norm or "LC 214" in col_norm:
                        val_excel = None
                        for k, v in excel_row.items():
                            if "ART" in normaliza(k) or "LC 214" in normaliza(k):
                                val_excel = v; break
                        texto = str(val_excel).strip() if not vazio(val_excel) else "(vazio)"
                        row_valores.append(texto)
                        row_tags.append("info")
                        info_cnt += 1
                        self.results.append(dict(arquivo=fname, campo=col,
                                                 esperado=texto, encontrado="—",
                                                 status="Referencia legal"))
                        continue
                    else:
                        val_excel = None
                        for k, v in excel_row.items():
                            if normaliza(k) == col_norm:
                                val_excel = v; break

                    # valor do XML
                    val_xml = self._buscar_tag(xml_flat, tags_xml) if tags_xml else None

                    ex_v = vazio(val_excel)
                    xm_v = vazio(val_xml)

                    # status
                    if col_norm == "CNPJ" and not ex_v and not xm_v:
                        status = "OK (raiz)" if cnpj_raiz(val_excel) == cnpj_raiz(val_xml) else "Divergente"
                    elif ex_v and xm_v:
                        status = "Campo vazio"
                    elif ex_v:
                        status = "Campo vazio no Excel"
                    elif xm_v:
                        status = "Campo vazio no XML"
                    elif normaliza(val_xml) == normaliza(val_excel):
                        status = "OK"
                    else:
                        status = "Divergente"

                    # contadores
                    if "OK" in status:            ok += 1
                    elif status == "Divergente":  err += 1
                    elif "no XML" in status:      vz_xml += 1
                    elif "no Excel" in status:    vz_excel += 1
                    elif status == "Campo vazio": vz_ambos += 1

                    tag = STATUS_TAG.get(status, "misto")
                    row_valores.append(status)
                    row_tags.append(tag)
                    self.results.append(dict(arquivo=fname, campo=col,
                                             esperado=str(val_excel) if not ex_v else "(vazio)",
                                             encontrado=str(val_xml) if not xm_v else "(vazio)",
                                             status=status))

                # cor da linha = pior status presente
                row_tag = "misto"
                for p in PRIORIDADE:
                    if p in row_tags:
                        row_tag = p
                        break

                self.tree.insert("", "end", values=row_valores, tags=(row_tag,))

            except Exception as e:
                vals = [fname] + [f"Erro: {e}"] + [""] * (len(COLUNAS) - 1)
                self.tree.insert("", "end", values=vals, tags=("erro",))
                err += 1

        total = ok + err + vz_xml + vz_excel + vz_ambos + info_cnt
        self.status_var.set(
            f"Validacao concluida  —  {ok} OK  |  {err} Divergentes  |  "
            f"{vz_xml} Vazio XML  |  {vz_excel} Vazio Excel  |  "
            f"{vz_ambos} Vazio Ambos  |  {info_cnt} Referencias  "
            f"[{len(self.xml_paths)} arquivo(s)]")

    # ── Excel ─────────────────────────────────────────────────────────────────
    def _read_excel(self, path):
        wb = openpyxl.load_workbook(path, data_only=True)
        cnpj_excel = ""
        itens      = []
        for ws in wb.worksheets:
            for tbl in ws.tables.values():
                nome = tbl.displayName or tbl.name or ""
                if nome == "TabelaResumoCadastro":
                    dados = self._ler_tabela(ws, tbl)
                    if dados:
                        for k, v in dados[0].items():
                            if normaliza(k) == "CNPJ":
                                cnpj_excel = str(v).strip() if not vazio(v) else ""
                                break
                elif nome == "Tabela2":
                    itens = self._ler_tabela(ws, tbl)
        wb.close()
        return cnpj_excel, itens

    def _ler_tabela(self, ws, tbl):
        rng     = ws[tbl.ref]
        rows    = [[cell.value for cell in row] for row in rng]
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        result  = []
        for row in rows[1:]:
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            d = {h: v for h, v in zip(headers, row) if h}
            if d:
                result.append(d)
        return result

    # ── XML ───────────────────────────────────────────────────────────────────
    def _read_xml_flat(self, path):
        tree = ET.parse(path)
        root = tree.getroot()
        data = {}
        self._flatten(root, "", data)
        return data

    def _flatten(self, element, prefix, data):
        tag = element.tag.split("}")[-1] if "}" in element.tag else element.tag
        key = f"{prefix}/{tag}" if prefix else tag
        for an, av in element.attrib.items():
            data[f"{key}@{an}"] = av
        text = (element.text or "").strip()
        if text:
            data[key] = text
            if tag not in data:
                data[tag] = text
        for child in element:
            self._flatten(child, key, data)

    def _buscar_tag(self, xml_flat, tags):
        for t in tags:
            if t in xml_flat:
                return xml_flat[t]
            tag_name = t.split("/")[-1]
            if tag_name in xml_flat:
                return xml_flat[tag_name]
            for k, v in xml_flat.items():
                if k.endswith(f"/{tag_name}") or k == tag_name:
                    return v
        return None

    # ── Exportar ──────────────────────────────────────────────────────────────
    def _export(self):
        if not self.results:
            messagebox.showwarning("Atencao", "Execute a validacao antes de exportar.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
            title="Salvar relatorio")
        if not path:
            return
        try:
            if path.endswith(".csv"):
                self._export_csv(path)
            else:
                self._export_excel(path)
            messagebox.showinfo("Exportado", f"Relatorio salvo em:\n{path}")
            self.status_var.set(f"Relatorio exportado: {os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("Erro ao exportar", str(e))

    def _export_excel(self, path):
        from openpyxl.styles import PatternFill, Font, Alignment
        from collections import defaultdict
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultado"
        ws.append(["Arquivo XML"] + COLUNAS)
        hdr_fill = PatternFill("solid", fgColor="7C6AF7")
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        fill_map = {
            "OK":                   PatternFill("solid", fgColor="D6F5E3"),
            "OK (raiz)":            PatternFill("solid", fgColor="D6F5E3"),
            "Divergente":           PatternFill("solid", fgColor="FAD7D7"),
            "Campo vazio no XML":   PatternFill("solid", fgColor="FFF0CC"),
            "Campo vazio no Excel": PatternFill("solid", fgColor="D0E8FF"),
            "Campo vazio":          PatternFill("solid", fgColor="E0E0E0"),
            "Referencia legal":     PatternFill("solid", fgColor="EFEFFA"),
        }

        por_arquivo = defaultdict(dict)
        for r in self.results:
            por_arquivo[r["arquivo"]][r["campo"]] = r["status"]

        for arquivo, campos in por_arquivo.items():
            ws.append([arquivo] + [campos.get(c, "") for c in COLUNAS])
            row_idx = ws.max_row
            for col_idx, col in enumerate(COLUNAS, start=2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill_map.get(campos.get(col, ""), PatternFill())

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22
        wb.save(path)

    def _export_csv(self, path):
        import csv
        from collections import defaultdict
        por_arquivo = defaultdict(dict)
        for r in self.results:
            por_arquivo[r["arquivo"]][r["campo"]] = r["status"]
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["Arquivo XML"] + COLUNAS)
            for arquivo, campos in por_arquivo.items():
                writer.writerow([arquivo] + [campos.get(c, "") for c in COLUNAS])


if __name__ == "__main__":
    app = XMLValidatorApp()
    app.mainloop()
