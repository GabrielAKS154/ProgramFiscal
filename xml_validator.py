import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import xml.etree.ElementTree as ET
import openpyxl
import os
import re
import unicodedata

BG      = "#1e1e2e"
SURFACE = "#2a2a3d"
ACCENT  = "#7c6af7"
ACCENT2 = "#5a4fcf"
TEXT    = "#e0e0f0"
MUTED   = "#8888aa"
SUCCESS = "#4caf82"
ERROR   = "#f06060"
WARNING = "#f0b060"
BORDER  = "#3a3a55"
INFO    = "#6ab0f5"
GRAY    = "#aaaaaa"

# Mapeamento Tabela2: coluna Excel → tags XML
MAPA_ITENS = {
    "TIPO":                    ["det/prod/xProd", "det/serv/xDescServ", "xProd", "xDescServ"],
    "NCM/NBS":                 ["NCM", "NBS", "det/prod/NCM", "det/serv/NBS"],
    "NCM":                     ["NCM", "det/prod/NCM"],
    "NBS":                     ["NBS", "det/serv/NBS"],
    "DESCRICAO":               ["xProd", "xDescServ", "det/prod/xProd", "det/serv/xDescServ"],
    "CST":                     ["CST", "CSOSN",
                                "det/imposto/ICMS/ICMS00/CST",
                                "det/imposto/ICMS/ICMSSN102/CSOSN",
                                "det/imposto/PIS/PISAliq/CST",
                                "det/imposto/COFINS/COFINSAliq/CST"],
    "CCLASSTRIB":              ["cClassTrib", "det/imposto/cClassTrib"],
    "CCREDPRES":               ["cCredPres",  "det/imposto/cCredPres"],
    "CCLASSTRIB OU CCREDPRES": ["cClassTrib", "cCredPres",
                                "det/imposto/cClassTrib", "det/imposto/cCredPres"],
    "ART. LC 214/2025":        [],  # apenas referencia legal
}

def normaliza_cnpj(v):
    return re.sub(r"\D", "", str(v or ""))

def cnpj_raiz(v):
    return normaliza_cnpj(v)[:8]

def normaliza(v):
    s = str(v or "").strip().upper()
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")

def vazio(v):
    return v is None or str(v).strip() == ""


class XMLValidatorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("XML Validator  v2.4")
        self.geometry("1200x720")
        self.minsize(900, 580)
        self.configure(bg=BG)
        self.excel_path       = tk.StringVar(value="Nenhum arquivo selecionado")
        self._excel_full_path = None
        self.xml_paths        = []
        self.results          = []
        self._build_ui()

    # ── UI ────────────────────────────────────────────────────────────────────
    def _build_ui(self):
        hdr = tk.Frame(self, bg=ACCENT, height=56)
        hdr.pack(fill="x")
        tk.Label(hdr, text="XML Validator", font=("Segoe UI", 16, "bold"),
                 bg=ACCENT, fg="white").pack(side="left", padx=20, pady=12)
        tk.Label(hdr, text="Validacao Fiscal NF-e  |  Reforma Tributaria  |  100% offline",
                 font=("Segoe UI", 9), bg=ACCENT, fg="#d0ccff").pack(side="left")

        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=20, pady=16)

        left = tk.Frame(body, bg=BG, width=300)
        left.pack(side="left", fill="y", padx=(0, 16))
        left.pack_propagate(False)

        self._section(left, "1. Formulario Excel (gabarito)")
        self._btn(left, "Selecionar Excel", self._load_excel).pack(fill="x", pady=(0, 4))
        tk.Label(left, textvariable=self.excel_path, bg=BG, fg=MUTED,
                 font=("Segoe UI", 8), wraplength=270, justify="left").pack(anchor="w")

        self._sep(left)
        self._section(left, "2. Arquivos XML (notas fiscais)")
        self._btn(left, "Adicionar XMLs", self._load_xmls).pack(fill="x", pady=(0, 4))
        self.xml_listbox = tk.Listbox(left, bg=SURFACE, fg=TEXT, selectbackground=ACCENT,
                                      relief="flat", bd=0, font=("Segoe UI", 8),
                                      height=7, highlightthickness=1,
                                      highlightbackground=BORDER)
        self.xml_listbox.pack(fill="x", pady=(0, 4))
        self._btn(left, "Remover selecionado", self._remove_xml, color=SURFACE).pack(fill="x")

        self._sep(left)
        self._section(left, "3. Executar")
        self._btn(left, "Validar XMLs", self._run, color=ACCENT).pack(fill="x", pady=(0, 4))
        self._btn(left, "Exportar relatorio", self._export, color=ACCENT2).pack(fill="x", pady=(0, 4))
        self._btn(left, "Limpar tudo", self._clear, color=SURFACE).pack(fill="x")

        self._sep(left)
        self._section(left, "Legenda")
        for cor, txt in [
            (SUCCESS, "OK"),
            (ERROR,   "Divergente"),
            (WARNING, "Campo vazio no XML"),
            (INFO,    "Campo vazio no Excel"),
            (GRAY,    "Campo vazio (ambos)"),
            (MUTED,   "Referencia legal"),
        ]:
            row = tk.Frame(left, bg=BG)
            row.pack(anchor="w", pady=1)
            tk.Frame(row, bg=cor, width=12, height=12).pack(side="left", padx=(0, 6))
            tk.Label(row, text=txt, bg=BG, fg=TEXT, font=("Segoe UI", 8)).pack(side="left")

        right = tk.Frame(body, bg=BG)
        right.pack(side="left", fill="both", expand=True)

        self.summary_frame = tk.Frame(right, bg=SURFACE, pady=10)
        self.summary_frame.pack(fill="x", pady=(0, 10))
        for label, attr, color in [
            ("Total",          "lbl_total",       TEXT),
            ("OK",             "lbl_ok",          SUCCESS),
            ("Divergente",     "lbl_err",         ERROR),
            ("Vazio XML",      "lbl_vazio_xml",   WARNING),
            ("Vazio Excel",    "lbl_vazio_excel", INFO),
            ("Vazio Ambos",    "lbl_vazio_ambos", GRAY),
            ("Referencia",     "lbl_info",        MUTED),
        ]:
            f = tk.Frame(self.summary_frame, bg=SURFACE)
            f.pack(side="left", expand=True)
            lv = tk.Label(f, text="0", font=("Segoe UI", 16, "bold"), bg=SURFACE, fg=color)
            lv.pack()
            tk.Label(f, text=label, font=("Segoe UI", 7), bg=SURFACE, fg=MUTED).pack()
            setattr(self, attr, lv)

        cols = ("arquivo", "campo", "esperado", "encontrado", "status")
        self.tree = ttk.Treeview(right, columns=cols, show="headings", height=22)
        for c, h, w in zip(cols,
                           ("Arquivo XML", "Campo",
                            "Esperado (Excel)", "Encontrado (XML)", "Status"),
                           (200, 200, 180, 180, 170)):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, minwidth=50, anchor="w")

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background=SURFACE, foreground=TEXT,
                        fieldbackground=SURFACE, rowheight=26, font=("Segoe UI", 9))
        style.configure("Treeview.Heading", background=BORDER, foreground=TEXT,
                        font=("Segoe UI", 9, "bold"))
        style.map("Treeview", background=[("selected", ACCENT)])
        self.tree.tag_configure("ok",          foreground=SUCCESS)
        self.tree.tag_configure("erro",        foreground=ERROR)
        self.tree.tag_configure("vazio_xml",   foreground=WARNING)
        self.tree.tag_configure("vazio_excel", foreground=INFO)
        self.tree.tag_configure("vazio_ambos", foreground=GRAY)
        self.tree.tag_configure("info",        foreground=MUTED)

        sb = ttk.Scrollbar(right, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="left", fill="y")

        self.status_var = tk.StringVar(value="Pronto. Selecione o Excel e os XMLs para comecar.")
        tk.Label(self, textvariable=self.status_var, bg=SURFACE, fg=MUTED,
                 font=("Segoe UI", 8), anchor="w", padx=12, pady=6).pack(fill="x", side="bottom")

    def _section(self, parent, text):
        tk.Label(parent, text=text, bg=BG, fg=ACCENT,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(10, 2))

    def _sep(self, parent):
        tk.Frame(parent, bg=BORDER, height=1).pack(fill="x", pady=8)

    def _btn(self, parent, text, cmd, color=ACCENT):
        return tk.Button(parent, text=text, command=cmd, bg=color, fg="white",
                         activebackground=ACCENT2, activeforeground="white",
                         relief="flat", bd=0, font=("Segoe UI", 9),
                         cursor="hand2", pady=8)

    # ── Carregar arquivos ─────────────────────────────────────────────────────
    def _load_excel(self):
        path = filedialog.askopenfilename(
            title="Selecionar formulario Excel",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")])
        if path:
            self.excel_path.set(os.path.basename(path))
            self._excel_full_path = path
            self.status_var.set(f"Excel carregado: {os.path.basename(path)}")

    def _load_xmls(self):
        paths = filedialog.askopenfilenames(
            title="Selecionar arquivos XML",
            filetypes=[("XML", "*.xml"), ("Todos", "*.*")])
        for p in paths:
            if p not in self.xml_paths:
                self.xml_paths.append(p)
                self.xml_listbox.insert("end", os.path.basename(p))
        self.status_var.set(f"{len(self.xml_paths)} arquivo(s) XML carregado(s).")

    def _remove_xml(self):
        sel = self.xml_listbox.curselection()
        if sel:
            idx = sel[0]
            self.xml_listbox.delete(idx)
            self.xml_paths.pop(idx)

    def _clear(self):
        self.excel_path.set("Nenhum arquivo selecionado")
        self._excel_full_path = None
        self.xml_paths.clear()
        self.xml_listbox.delete(0, "end")
        self.tree.delete(*self.tree.get_children())
        self.results.clear()
        self._update_summary()
        self.status_var.set("Limpo. Pronto para nova validacao.")

    # ── Validacao principal ───────────────────────────────────────────────────
    def _run(self):
        try:
            self._run_inner()
        except Exception as e:
            import traceback
            messagebox.showerror("Erro inesperado", traceback.format_exc())

    def _run_inner(self):
        if not self._excel_full_path:
            messagebox.showwarning("Atencao", "Selecione o formulario Excel primeiro.")
            return
        if not self.xml_paths:
            messagebox.showwarning("Atencao", "Adicione pelo menos um arquivo XML.")
            return

        self.tree.delete(*self.tree.get_children())
        self.results.clear()
        self.status_var.set("Processando...")
        self.update_idletasks()

        try:
            cnpj_excel, itens = self._read_excel(self._excel_full_path)
        except Exception as e:
            messagebox.showerror("Erro ao ler Excel", str(e))
            return

        ok = err = vazio_xml = vazio_excel = vazio_ambos = info = 0

        for xml_path in self.xml_paths:
            fname = os.path.basename(xml_path)
            try:
                xml_flat = self._read_xml_flat(xml_path)

                # ── 1. CNPJ da TabelaResumoCadastro ───────────────────────
                cnpj_xml = self._buscar_tag(xml_flat, ["emit/CNPJ", "CNPJ"])
                ex_v = vazio(cnpj_excel)
                xm_v = vazio(cnpj_xml)

                if ex_v and xm_v:
                    status, tag = "Campo vazio", "vazio_ambos"; vazio_ambos += 1
                elif ex_v:
                    status, tag = "Campo vazio no Excel", "vazio_excel"; vazio_excel += 1
                elif xm_v:
                    status, tag = "Campo vazio no XML", "vazio_xml"; vazio_xml += 1
                elif cnpj_raiz(cnpj_excel) == cnpj_raiz(cnpj_xml):
                    status, tag = "OK (raiz)", "ok"; ok += 1
                else:
                    status, tag = "Divergente", "erro"; err += 1

                self._add_row(fname, "CNPJ",
                              cnpj_excel or "(vazio)",
                              cnpj_xml   or "(vazio)",
                              status, tag)

                # ── 2. Todas as colunas da Tabela2 ────────────────────────
                for row in itens:
                    for col, valor_excel in row.items():
                        col_norm = normaliza(col)

                        # Art. LC 214/2025 — apenas referencia, nao compara
                        if "ART." in col_norm or "LC 214" in col_norm:
                            if not vazio(valor_excel):
                                self._add_row(fname, col,
                                              str(valor_excel), "—",
                                              "Referencia legal", "info")
                                info += 1
                            continue

                        tags      = self._resolver_tags(col_norm)
                        valor_xml = self._buscar_tag(xml_flat, tags) if tags else None

                        ex_v = vazio(valor_excel)
                        xm_v = vazio(valor_xml)

                        if ex_v and xm_v:
                            status, tag = "Campo vazio", "vazio_ambos"
                            vazio_ambos += 1
                        elif ex_v:
                            status, tag = "Campo vazio no Excel", "vazio_excel"
                            vazio_excel += 1
                        elif xm_v:
                            status, tag = "Campo vazio no XML", "vazio_xml"
                            vazio_xml += 1
                        elif normaliza(valor_xml) == normaliza(valor_excel):
                            status, tag = "OK", "ok"
                            ok += 1
                        else:
                            status, tag = "Divergente", "erro"
                            err += 1

                        self._add_row(fname, col,
                                      str(valor_excel) if not ex_v else "(vazio)",
                                      str(valor_xml)   if not xm_v else "(vazio)",
                                      status, tag)

            except Exception as e:
                self._add_row(fname, "-", "-", f"Erro: {e}", "Falha", "erro")
                err += 1

        total = ok + err + vazio_xml + vazio_excel + vazio_ambos + info
        self._update_summary(total, ok, err, vazio_xml, vazio_excel, vazio_ambos, info)
        self.status_var.set(
            f"Validacao concluida  —  {ok} OK  |  {err} Divergentes  |  "
            f"{vazio_xml} Vazio XML  |  {vazio_excel} Vazio Excel  |  "
            f"{vazio_ambos} Vazio Ambos  |  {info} Referencias")

    def _add_row(self, arquivo, campo, esperado, encontrado, status, tag):
        self.tree.insert("", "end",
                         values=(arquivo, campo, esperado, encontrado, status),
                         tags=(tag,))
        self.results.append(dict(arquivo=arquivo, campo=campo,
                                 esperado=esperado, encontrado=encontrado,
                                 status=status))

    # ── Leitura do Excel ──────────────────────────────────────────────────────
    def _read_excel(self, path):
        wb = openpyxl.load_workbook(path, data_only=True)
        cnpj_excel = ""
        itens      = []

        for ws in wb.worksheets:
            for tbl in ws.tables.values():
                nome = tbl.displayName or tbl.name or ""
                if nome == "TabelaResumoCadastro":
                    dados = self._ler_tabela(ws, tbl)
                    if dados:
                        for k, v in dados[0].items():
                            if normaliza(k) == "CNPJ":
                                cnpj_excel = str(v).strip() if not vazio(v) else ""
                                break
                elif nome == "Tabela2":
                    itens = self._ler_tabela(ws, tbl)

        wb.close()
        return cnpj_excel, itens

    def _ler_tabela(self, ws, tbl):
        rng     = ws[tbl.ref]
        rows    = [[cell.value for cell in row] for row in rng]
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        result  = []
        for row in rows[1:]:
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            d = {h: v for h, v in zip(headers, row) if h}
            if d:
                result.append(d)
        return result

    # ── Leitura do XML ────────────────────────────────────────────────────────
    def _read_xml_flat(self, path):
        tree = ET.parse(path)
        root = tree.getroot()
        data = {}
        self._flatten(root, "", data)
        return data

    def _flatten(self, element, prefix, data):
        tag = element.tag.split("}")[-1] if "}" in element.tag else element.tag
        key = f"{prefix}/{tag}" if prefix else tag
        for an, av in element.attrib.items():
            data[f"{key}@{an}"] = av
        text = (element.text or "").strip()
        if text:
            data[key] = text
            if tag not in data:
                data[tag] = text
        for child in element:
            self._flatten(child, key, data)

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _resolver_tags(self, col_norm):
        for k, v in MAPA_ITENS.items():
            if normaliza(k) == col_norm:
                return v
        return [col_norm]

    def _buscar_tag(self, xml_flat, tags):
        for t in tags:
            if t in xml_flat:
                return xml_flat[t]
            tag_name = t.split("/")[-1]
            if tag_name in xml_flat:
                return xml_flat[tag_name]
            for k, v in xml_flat.items():
                if k.endswith(f"/{tag_name}") or k == tag_name:
                    return v
        return None

    # ── Resumo ────────────────────────────────────────────────────────────────
    def _update_summary(self, total=0, ok=0, err=0,
                        vazio_xml=0, vazio_excel=0, vazio_ambos=0, info=0):
        self.lbl_total.config(text=str(total))
        self.lbl_ok.config(text=str(ok))
        self.lbl_err.config(text=str(err))
        self.lbl_vazio_xml.config(text=str(vazio_xml))
        self.lbl_vazio_excel.config(text=str(vazio_excel))
        self.lbl_vazio_ambos.config(text=str(vazio_ambos))
        self.lbl_info.config(text=str(info))

    # ── Exportar ──────────────────────────────────────────────────────────────
    def _export(self):
        if not self.results:
            messagebox.showwarning("Atencao", "Execute a validacao antes de exportar.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
            title="Salvar relatorio")
        if not path:
            return
        try:
            if path.endswith(".csv"):
                self._export_csv(path)
            else:
                self._export_excel(path)
            messagebox.showinfo("Exportado", f"Relatorio salvo em:\n{path}")
            self.status_var.set(f"Relatorio exportado: {os.path.basename(path)}")
        except Exception as e:
            messagebox.showerror("Erro ao exportar", str(e))

    def _export_excel(self, path):
        from openpyxl.styles import PatternFill, Font, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultado"
        ws.append(["Arquivo XML", "Campo",
                   "Esperado (Excel)", "Encontrado (XML)", "Status"])
        hdr_fill = PatternFill("solid", fgColor="7C6AF7")
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        fill_ok          = PatternFill("solid", fgColor="D6F5E3")
        fill_err         = PatternFill("solid", fgColor="FAD7D7")
        fill_vazio_xml   = PatternFill("solid", fgColor="FFF0CC")
        fill_vazio_excel = PatternFill("solid", fgColor="D0E8FF")
        fill_vazio_ambos = PatternFill("solid", fgColor="E0E0E0")
        fill_info        = PatternFill("solid", fgColor="EFEFFA")
        for r in self.results:
            ws.append([r["arquivo"], r["campo"],
                       r["esperado"], r["encontrado"], r["status"]])
            s = r["status"]
            if "OK" in s:
                fill = fill_ok
            elif "Divergente" in s or "Falha" in s:
                fill = fill_err
            elif s == "Campo vazio":
                fill = fill_vazio_ambos
            elif "Campo vazio no XML" in s:
                fill = fill_vazio_xml
            elif "Campo vazio no Excel" in s:
                fill = fill_vazio_excel
            else:
                fill = fill_info
            for cell in ws[ws.max_row]:
                cell.fill = fill
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 28
        wb.save(path)

    def _export_csv(self, path):
        import csv
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(
                f, fieldnames=["arquivo", "campo",
                               "esperado", "encontrado", "status"])
            writer.writeheader()
            writer.writerows(self.results)


if __name__ == "__main__":
    app = XMLValidatorApp()
    app.mainloop()
